import flet as ft
from flet import *
import re
from fletsheet_formula import Formulas
import openpyxl
import sys 
import os
import random
import datetime


class TextFieldTable():
    """
    Una matriz de celdas que se puede editar y desplazar.
    """

    def __init__(self, 
                excel_file_path,
                page_width, 
                page_height,
                cell_height: OptionalNumber = 30,
                cell_width: OptionalNumber = 100,
                ):
        
        self.page_width = page_width
        self.page_height = page_height
        self.ROWS = int((page_height / 30)*0.8)
        self.COLS = int((page_width / 100)*0.9)
        self.selected_cells = [] #inicializar como lista vacía
        self.dragging = False
        self.double_clicked = False
        self.current_selected_cell = None
        self.editing_cell = None
        self.clipboard = [] #contenido copiado o cortado puede ser una lista
        self.double_clicked = False
        self.visible_start_row = 1
        self.visible_end_row =  self.ROWS
        self.visible_start_col = 1
        self.visible_end_col = self.COLS
        self.start_cell = None 
        self.table_rows = []
        self.table_initialized = False  # Inicializa el estado de la tabla
        self.cell_height = cell_height  
        self.cell_width = cell_width
        self.btn_hoja = False
        self.edited_cells = {}  
        self.undo_stack = []  # Pila para deshacer cambios
        self.excel_file_path = excel_file_path
        self.excel_data = self.load_excel_data(self.excel_file_path)
        self.current_sheet = next(iter(self.excel_data), None)
        self.cells =  [[None for _ in range(self.COLS)] for _ in range(self.ROWS)]  # Matriz de celdas
        #self.cells = {sheet_name: [[None for _ in range(self.COLS)] for _ in range(self.ROWS)] for sheet_name in self.excel_data} #organizar las celdas de acuerdo a hojas, para más adelante
        self.text_size = 14
        self.is_writing_formula = False
        self.cell_colors = {}  # Diccionario para almacenar los colores de las celdas
        self.formula_container = None
        self.container_row = None
        self.container_col = None



    def is_android_device(self, page):
        self.page.platform == ft.PagePlatform.ANDROID
    
    def is_ios_device(self, page):
        self.page.platform == ft.PagePlatform.IOS
    
    def is_pc_device(self, page):
        self.page.platform == ft.PagePlatform.MACOS or self.page.platform == ft.PagePlatform.LINUX or self.page.platform == ft.PagePlatform.WINDOWS
        

    def is_packege_device(self, page):
        # verificar si es un dispositivo empaquetado
        if getattr(sys, 'frozen', False):
            # En un entorno empaquetado
            return True
        else:
            # En un entorno de desarrollo
            return False
    
 
    def load_excel_data(self, filepath):
        print("se está cargando data excel")
        workbook = openpyxl.load_workbook(filepath, data_only=False)  # Cambiar data_only a False
        data = {}

        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            sheet_data = []
            for row in worksheet.iter_rows():
                row_data = []
                for cell in row:
                    if cell.data_type == 'f':  # Verificar si la celda contiene una fórmula
                        cell_value = cell.value  # Preceder la fórmula con '='
                    else:
                        cell_value = cell.value  # Obtener el valor de la celda como es
                    row_data.append(cell_value)
                sheet_data.append(row_data)
            data[sheet_name] = sheet_data

        return data

    def create_formula_container(self, page):
        self.formula_container = ft.Container(
            content=ft.Text(""),
            width=page.width*0.9,
            height=page.height*0.05,
            bgcolor=ft.colors.GREY_200,
            on_click=lambda e: self.on_container_click(page)
        )
    
    def iluminar(self, cell, page):
        cell_key = (cell.row, cell.col)
        if cell_key not in self.cell_colors:
            # Generar valores aleatorios solo si la celda no ha sido resaltada antes
            r = random.randint(0, 255)
            g = random.randint(0, 255)
            b = random.randint(0, 255)
            color_aleatorio = f'#ff{r:02x}{g:02x}{b:02x}'
            self.cell_colors[cell_key] = color_aleatorio
        else:
            color_aleatorio = self.cell_colors[cell_key]

        # Aplicar el color al borde de la celda
        cell.border = ft.border.all(1.5, color_aleatorio)
            

    
    def on_container_click(self, page):
        # Crear y configurar CupertinoTextField para la edición
        text_field = CupertinoTextField(
            value= self.formula_container.content.value,
            on_submit=lambda  e: save_container_edited_value(self.container_row, self.container_col, self.formula_container.content.value, page),
            on_change=lambda e: on_container_textfield_change(e, self.formula_container, page),
            autofocus=True,
            placeholder_text="",
            text_size= self.text_size,
        )

        # Actualizar el contenido de la celda para mostrar el TextField
        self.formula_container.content = text_field
        page.update()

        def on_container_textfield_change(e, cell, page):
            # Verifica si el texto comienza con '='

            if self.formula_container.content.value.startswith("="):
                self.is_writing_formula = True
            
            if self.is_writing_formula:
                cell_references = parse_cell_container_references(cell.content.value)
                for ref in cell_references:
                    cell_to_highlight = get_cell_container_from_reference(ref)
                    self.iluminar(cell_to_highlight, page)
            else:
                self.is_writing_formula = False
                # Elimina el resaltado si es necesario
                # ...
        
        def parse_cell_container_references(formula):
            # Regular expression to find cell references like A1, B2, etc.
            pattern = r'([A-Za-z]+[0-9]+)'
            return re.findall(pattern, formula)
       
        
        def get_cell_container_from_reference(ref):
            # Assuming the cell references are in the format 'A1', 'B2', etc.
            # Convert column letter to column index (e.g., 'A' -> 0, 'B' -> 1)
            col = ord(ref[0].upper()) - ord('A')
            # Convert row number to row index (e.g., '1' -> 0, '2' -> 1)
            row = int(ref[1:]) - 1
            # Return the cell object at the calculated row and column
            return self.cells[row][col]

        
        def save_container_edited_value(row, col, value, page):
            # Actualizar el valor de la celda en la estructura de datos
            self.excel_data[self.current_sheet][row][col] = value
            previous_value = self.formula_container.content.value
            current_text = self.formula_container.content.value

            # Actualizar la visualización de la celda para mostrar el nuevo valor
            cell = self.cells[row][col]
            cell.content = ft.Text(value, size=self.text_size)  # Reemplazar el TextField por un Text con el nuevo valor

            self.double_clicked = False 

            
            if cell.content.value.startswith("="):
                        row, col = row, col
                        formula = cell.content.value  
                        result = Formulas().evaluate_formula(formula, "withcell", cells=self.cells,) 
                        cell.formula = formula  
                        cell.content.value = str(result) 
                        self.unhighlight_cell_colors(page)

            self.editing_cell = None  

            almacenar_datoescrito(row, col, current_text, previous_value)

            page.update()

            
        

        def almacenar_datoescrito(current_row, current_col, current_cell, previous_value):
            adjusted_row = current_row - 1 + self.visible_start_row
            adjusted_col = current_col - 1 + self.visible_start_col

            sheet_name = self.current_sheet
            if sheet_name not in self.edited_cells:
                self.edited_cells[sheet_name] = {}

            # Determinar si el contenido de la celda es una fórmula o un valor
            if isinstance(self.formula_container.content.value, str) and self.formula_container.content.value.startswith('='):
                # Almacenar como fórmula
                edited_value = self.formula_container.content.value
            else:
                # Almacenar como valor normal
                edited_value = self.formula_container.content.value

            self.edited_cells[sheet_name][(adjusted_row, adjusted_col)] = edited_value

            self.undo_stack.append(('edit', sheet_name, current_row, current_col, previous_value, edited_value))
            print(f"edited_cells:{self.edited_cells}" )
        


    def highlight_cell(self, cell, page):
        
        if cell not in self.selected_cells:
            cell.border = ft.border.all(1.5, ft.colors.PINK_400)
            self.selected_cells.append(cell)
            print(f"Resaltando celda en {cell.row}, {cell.col}")
            page.update()
    
    

    def unhighlight_cell(self, cell, page):
        if cell in self.selected_cells:
            cell.border = ft.border.all(0.3, ft.colors.GREEN_500)
            self.selected_cells.remove(cell)
            print(f"desresaltando celda en {cell.row}, {cell.col}")
            page.update()

    
    def unhighlight_cell_colors(self, page):
        """
        Desmarca las celdas seleccionadas y limpia sus colores.
        """
        if self.cell_colors:
            for cell_key in self.cell_colors.keys():
                # cell_key contiene la tupla (row, column)
                row, column = cell_key
                # Asumiendo que las celdas están organizadas en una matriz 'self.cells'
                cell = self.cells[row][column]

                cell.border = ft.border.all(0.3, ft.colors.GREEN_500)
                if cell in self.selected_cells:
                    self.selected_cells.remove(cell)
            self.cell_colors.clear()
            page.update()

        
        
        
            

    def on_keyboard_event(self, e:ft.KeyboardEvent, page):

        
        def parse_cell_references(formula):
            # Regular expression to find cell references like A1, B2, etc.
            pattern = r'([A-Za-z]+[0-9]+)'
            return re.findall(pattern, formula)
       
        
        def get_cell_from_reference(ref):
            # Assuming the cell references are in the format 'A1', 'B2', etc.
            # Convert column letter to column index (e.g., 'A' -> 0, 'B' -> 1)
            col = ord(ref[0].upper()) - ord('A')
            # Convert row number to row index (e.g., '1' -> 0, '2' -> 1)
            row = int(ref[1:]) - 1
            # Return the cell object at the calculated row and column
            
            return self.cells[row][col]

        def almacenar_datoescrito(current_row, current_col, current_cell, previous_value):
            adjusted_row = current_row - 1 + self.visible_start_row
            adjusted_col = current_col - 1 + self.visible_start_col

            sheet_name = self.current_sheet
            if sheet_name not in self.edited_cells:
                self.edited_cells[sheet_name] = {}

            # Determinar si el contenido de la celda es una fórmula o un valor
            if isinstance(current_cell.content.value, str) and current_cell.content.value.startswith('='):
                # Almacenar como fórmula
                edited_value = current_cell.content.value
            else:
                # Almacenar como valor normal
                edited_value = current_cell.content.value

            self.edited_cells[sheet_name][(adjusted_row, adjusted_col)] = edited_value

            self.undo_stack.append(('edit', sheet_name, current_row, current_col, previous_value, edited_value))
            print(f"edited_cells:{self.edited_cells}" )
           
            
        def find_last_value_cell(direction):
            if direction == "down":
                for row in range(len(self.cells) - 1, -1, -1):
                    for col in range(len(self.cells[row])):
                        if self.cells[row][col].content.value:
                            print(f"row{row}, col{col}")  # Solo para depuración
                            return row, col
            elif direction == "right":
                for col in range(len(self.cells[0]) - 1, -1, -1):
                    for row in range(len(self.cells)):
                        if self.cells[row][col].content.value:
                            print(f"row{row}, col{col}")  # Solo para depuración
                            return row, col
            return None, None  # Retorna None si no se encuentra ninguna celda

        # Verificar si hay alguna celda seleccionada
        if not self.selected_cells:
            return
        
        # Utilizar la última celda seleccionada
        current_cell = self.selected_cells[-1]

        current_row = current_cell.row
        current_col = current_cell.col
        
      
        
          

        if re.match(r'^[a-zA-Z0-9\s=+\-*/()!@#$%^&*<>?{}\[\]~`|:;"\',.<>\/\^_`{|}~\\]$', e.key):
            if not e.ctrl:
                previous_value = current_cell.content.value  # Capturar el valor original

                if self.editing_cell != current_cell and not self.double_clicked:
                    current_cell.content.value = ""  # Borra el contenido existente
                    self.editing_cell = current_cell  # Actualiza el estado de edición

                current_text = current_cell.content.value #obtener texto actual del objeto Text

                # Si la tecla es alfanumérica, considera mayúsculas y minúsculas
                if re.match(r'^[a-zA-Z0-9=+\-*/()!@#$%^&*<>?{}[\]~`|]$', e.key):
                    if e.shift:
                        # Añadir una comprobación adicional para asegurar que e.key es una letra
                        if re.match(r'^[a-zA-Z]$', e.key):
                            current_cell.content.value = current_text + e.key.upper()

                        else: 
                            current_cell.content.value = current_text + e.key
                        
                    else:
                        current_cell.content.value = current_text + e.key.lower()
                else:  # Para otros caracteres como '=', '+', '-', etc.
                        current_cell.content.value = current_text + e.key
                        
                almacenar_datoescrito(current_row, current_col, current_cell, previous_value)

            page.update()
        
        
        if e.key == "=":
            self.is_writing_formula = True

        if self.is_writing_formula:
            cell_references = parse_cell_references(current_cell.content.value)
            for ref in cell_references:
                cell_to_highlight = get_cell_from_reference(ref)
                self.iluminar(cell_to_highlight, page)

        if e.key == "Enter":
        
            self.double_clicked = False 

            if not self.selected_cells:
                return

            if self.double_clicked:
                return

            if self.editing_cell:  
                if self.editing_cell.content.value.startswith("="):
                    row, col = self.editing_cell.row, self.editing_cell.col
                    formula = self.editing_cell.content.value  
                    result = Formulas().evaluate_formula(formula, "withcell", cell= self.cells) 
                    self.editing_cell.formula = formula  
                    self.editing_cell.content.value = str(result)  # Asegurarse de que el resultado se refleje en la celda
                    print(self.cell_colors)
                    self.unhighlight_cell_colors(page)


                self.editing_cell = None  
            page.update()

            return  # Finalizar el manejo del evento aquí, ya que hemos manejado la tecla "Enter"

        if e.ctrl == True:
            if e.key == "Arrow Down":
                last_row, _ = find_last_value_cell("down")
                if last_row:
                    self.visible_start_row = max(1, last_row - 12)
                    self.visible_end_row = min(self.ROWS, last_row + 1)
                    print(f"self.visible_start_row{self.visible_start_row}, self.visible_end_row:{self.visible_end_row}")
            elif e.key == "Arrow Right":
                _, last_col = find_last_value_cell("right")
                if last_col:
                    self.visible_start_col = max(1, last_col - 10)
                    self.visible_end_col = min(self.COLS, last_col + 1)
                    print(f"self.visible_start_col{self.visible_start_row}, self.visible_end_col:{self.visible_end_row}")

            # Actualiza las celdas visibles y la página
            self.update_visible_cells()
            self.update_indices()
            page.update()
                


            if e.key.lower() == "c":
                start_row, start_col = self.selected_cells[0].row, self.selected_cells[0].col
                end_row, end_col = self.selected_cells[-1].row, self.selected_cells[-1].col
                self.clipboard = []
                clipboard_str = ""  # Una cadena para almacenar el contenido en formato de texto plano
                for row in range(start_row, end_row+1):
                    row_values = []
                    row_str_values = []  # Almacena los valores de la fila como cadenas
                    for col in range(start_col, end_col+1):
                        value = self.cells[row][col].content.value
                        row_values.append(value)
                        row_str_values.append(str(value))
                    self.clipboard.append(row_values)
                    clipboard_str += "\t".join(row_str_values) + "\n"  # Separar las celdas con tabuladores y las filas con saltos de línea

                # Actualizar el portapapeles del sistema
                page.set_clipboard(clipboard_str)

                print(f"Contenido copiado: {self.clipboard}")

            elif e.key.lower() == "v":
                clipboard_content = page.get_clipboard()  # Obtener el contenido del portapapeles
                if clipboard_content:
                    # Dividir el contenido en filas y celdas
                    rows = clipboard_content.split("\n")
                    matrix = [row.split("\t") for row in rows]

                    # Pegar este contenido en tu programa
                    if self.selected_cells:  # Verificar que hay una celda seleccionada donde pegar
                        start_row, start_col = self.selected_cells[-1].row, self.selected_cells[-1].col
                        for i, row_values in enumerate(matrix):
                            for j, value in enumerate(row_values):
                                row = start_row + i
                                col = start_col + j
                                if 0 <= row < self.ROWS and 0 <= col < self.COLS:
                                    current_cell = self.cells[row][col]
                                    previous_value = current_cell.content.value
                                    current_cell.content.value = value
                                    almacenar_datoescrito(row, col, current_cell, previous_value)  # Actualizar self.edited_cells
                                    if value.startswith("="):  # Si es una fórmula
                                        Formulas().evaluate_formula(value, "withcell", cells=self.cells)  # Evaluar fórmulas         

                    page.update()

            elif e.key.lower() == "x":
                if self.selected_cells:
                    start_row, start_col = self.selected_cells[0].row, self.selected_cells[0].col
                    end_row, end_col = self.selected_cells[-1].row, self.selected_cells[-1].col
                    self.clipboard = []
                    clipboard_str = ""  # Una cadena para almacenar el contenido en formato de texto plano
                    for row in range(start_row, end_row + 1):
                        for col in range(start_col, end_col + 1):
                            current_cell = self.cells[row][col]
                            previous_value = current_cell.content.value  # Capturar el valor anterior
                            current_cell.content.value = ""
                            almacenar_datoescrito(row, col, current_cell, previous_value)
                        self.clipboard.append(row_values)
                        clipboard_str += "\t".join(row_str_values) + "\n"  # Separar las celdas con tabuladores y las filas con saltos de línea

                    # Actualizar el portapapeles del sistema
                    page.set_clipboard(clipboard_str)
                    #print(f"Contenido cortado: {self.clipboard}")
                    page.update()
                
            elif e.key.lower() == "z":
                 if self.undo_stack:
                    change_type, sheet_name, row, col, old_value, new_value = self.undo_stack.pop()
                    print(f"stack.pop: {self.undo_stack.pop()} ")
                    print(f"Deshaciendo: {change_type}, fila: {row}, columna: {col}, valor anterior: {old_value}, valor nuevo: {new_value}")
                    
                    # Actualizar la celda específica con el valor antiguo
                    self.cells[row][col].content.value = old_value

                    # Actualizar el diccionario de celdas editadas
                    sheet_name = self.current_sheet
                    if sheet_name not in self.edited_cells:
                        self.edited_cells[sheet_name] = {}
                    self.edited_cells[sheet_name][(row + 1, col + 1)] = old_value  # Asegurarse de que las coordenadas se ajusten a la estructura de edited_cells

                    page.update()


        # Luego, manejar las teclas de flecha
    
    
        scroll_needed = False
        mirar_scroll_needed = self.selected_cells[0]

        row_scroll_needed = mirar_scroll_needed.row
        col_scroll_needed = mirar_scroll_needed.col
       

        if not self.double_clicked:

            if e.shift and self.start_cell is None:
                # Si 'Shift' está presionado y no hay una celda de inicio establecida,
                # se establece la celda actual como la celda de inicio antes de moverse.
                self.start_cell = self.cells[current_row][current_col]
                print(f"Estableciendo start_cell a ({self.start_cell.row}, {self.start_cell.col}) al inicio de la selección")

            self.end_cell = self.cells[current_row][current_col]

              # La lógica para moverse entre las celdas se mantiene igual
            if e.key == "Arrow Up" and row_scroll_needed == self.visible_start_row:
                scroll_needed = True
                print("scroll arriba")
                print(f"acual row scorll needed: {row_scroll_needed} y actual col needed {col_scroll_needed} ")
                print( f"limite row {self.visible_start_row} limite col {self.visible_start_col}")
                self.update_showercontainer(page,current_row,current_col)

                current_row -= 1

            elif e.key == "Arrow Down" and row_scroll_needed == self.visible_end_row -1:
                scroll_needed = True
                print("Scroll Abajo")
                print(f"acual row scorll needed: {row_scroll_needed} y actual col needed {col_scroll_needed} ")
                print( f"limite row {self.visible_end_row} limite col {self.visible_end_col}")
                self.update_showercontainer(page,current_row,current_col)
                
                current_row += 1

            elif e.key == "Arrow Left" and col_scroll_needed == self.visible_start_col:
                scroll_needed = True
                print("scroll izquierda")
                print(f"acual row scorll needed: {row_scroll_needed} y actual col needed {col_scroll_needed} ")
                print( f"limite row {self.visible_start_row} limite col {self.visible_start_col}")
                self.update_showercontainer(page,current_row,current_col)
                current_col -= 1

            elif e.key == "Arrow Right" and col_scroll_needed == self.visible_end_col-1:
                scroll_needed = True
                print("scroll derecha")
                print(f"acual row scorll needed: {row_scroll_needed} y actual col needed {col_scroll_needed} ")
                print( f"limite row {self.visible_end_row} limite col {self.visible_end_col}")
                self.update_showercontainer(page,current_row,current_col)
                current_col += 1


            elif e.key == "Arrow Up" and current_row > 0:
                current_row -= 1
                self.update_showercontainer(page,current_row,current_col)
                

            elif e.key == "Arrow Down" and current_row < self.ROWS - 1:
                current_row += 1
                print(f"acual row scorll needed: {row_scroll_needed} y actual col needed {col_scroll_needed} ")
                print( f"limite row {self.visible_end_row} limite col {self.visible_end_col}")
                self.update_showercontainer(page,current_row,current_col)
                

            elif e.key == "Arrow Left" and current_col > 0:
                current_col -= 1
                self.update_showercontainer(page,current_row,current_col)
                

            elif e.key == "Arrow Right" and current_col < self.COLS - 1:
                current_col += 1
                self.update_showercontainer(page,current_row,current_col)
                
                   
            else:
                return
            
            if scroll_needed:

                cell = self.cells[row_scroll_needed][col_scroll_needed]
                self.end_cell = cell
                # Si se necesita desplazamiento, actualiza los índices de las filas/columnas visibles
                if e.key == "Arrow Up":
                    
                    
                    self.visible_end_row = max(self.visible_end_row - 1, self.ROWS)
                    self.visible_start_row =  max(self.visible_start_row -1, 0)


                elif e.key == "Arrow Down":

                    self.visible_end_row = self.ROWS
                    self.visible_start_row = self.visible_start_row +1


                elif e.key == "Arrow Left":
                

                    self.visible_start_col = max(self.visible_start_col - 1, 0)
                    self.visible_end_col = max(self.visible_end_col -1, self.COLS)

                    


                elif e.key == "Arrow Right":
                

                    self.visible_start_col = self.visible_start_col +1
                    self.visible_end_col = self.COLS


                # Actualiza las celdas visibles y la interfaz de usuario
                self.update_visible_cells()
                self.update_indices()
                page.update()


            # Si 'Shift' está presionado, resaltar la nueva celda
            if not scroll_needed:
                cell = self.cells[current_row][current_col]
                self.end_cell = cell
            
            

            if e.shift:
                
                if self.start_cell is None :
                    # Actualizar la celda inicial para la nueva selección
                    self.start_cell = cell
                    print(f"Estableciendo start_cell a ({self.start_cell.row}, {self.start_cell.col}) porque se presionó shift")
        
                start_row, start_col = self.start_cell.row , self.start_cell.col 
                end_row, end_col = self.end_cell.row, self.end_cell.col

                new_selected_cells = []
                for row in range(min(start_row, end_row), max(start_row, end_row) + 1):
                    for col in range(min(start_col, end_col), max(start_col, end_col) + 1):
                        new_selected_cells.append(self.cells[row][col])

                for cell in self.selected_cells:
                    if cell not in new_selected_cells:
                        self.unhighlight_cell(cell, page)

                for cell in new_selected_cells:
                    if cell not in self.selected_cells:
                        self.highlight_cell(cell, page)

                self.selected_cells = new_selected_cells

                print(f"Current row: {current_row}, Current col: {current_col}")
                print(f"Start row: {start_row}, Start col: {start_col}")
                print(f"End row: {end_row}, End col: {end_col}")

                if not e.shift:
                    self.clear_all_highlights(page)
                    self.highlight_cell(cell, page)
                    start_row = None
                    start_col = None
                    #print(f"la start cell es: {self.start_cell} ")
            else:
                self.clear_all_highlights(page)
                self.highlight_cell(cell, page)
                self.start_cell = None  # Reset the start cell
                #print(f"la start cell está en: {self.start_cell} ")

            page.update()
            
    def clear_all_highlights(self, page):
        for cell in self.selected_cells[:]:  # Haz una copia de la lista para iterar
            self.unhighlight_cell(cell, page)
        self.selected_cells.clear()  # Limpia la lista original
    
    def update_showercontainer(self, page, row, col):
        # Ajustar fila y columna basándose en el desplazamiento
        adjusted_row = row + self.visible_start_row - 1
        adjusted_col = col + self.visible_start_col - 1

        self.container_col = adjusted_col
        self.container_row = adjusted_row

        sheet_data = self.excel_data[self.current_sheet]  # Acceder a los datos de la hoja actual

        # Asegurarse de que la fila y la columna ajustadas están dentro de los límites
        if adjusted_row < len(sheet_data) and adjusted_col < len(sheet_data[adjusted_row]):
            cell_value = sheet_data[adjusted_row][adjusted_col]  # Obtener el valor de la celda

            # Actualizar el contenido del contenedor de fórmulas
            self.formula_container.content = ft.Text(cell_value)
            page.update()
        
    def on_single_click(self, e: ft.TapEvent, page):
        
        print("On single click")
    
        cell = self.cells[e.control.row][e.control.col]

        self.update_showercontainer(page, e.control.row, e.control.col)

        # Desresaltar todas las celdas previamente seleccionadas
        self.clear_all_highlights(page)

        # Resaltar la celda actualmente seleccionada
        self.highlight_cell(cell, page)

        self.unhighlight_cell_colors(page)

        # Actualizar la celda actualmente seleccionada
        self.current_selected_cell = cell

        self.table_initialized = True 
        
       

    
    def on_double_click(self, e: ft.TapEvent, page):
        self.double_clicked = True
        cell = self.cells[e.control.row][e.control.col]
        # Desresaltar todas las celdas previamente seleccionadas
        self.clear_all_highlights(page)

        # Resaltar la celda actualmente seleccionada
        self.highlight_cell(cell, page)

        # Actualizar la celda actualmente seleccionada
        self.current_selected_cell = cell

        row= e.control.row
        col = e.control.col

        self.table_initialized = True 
        
        # Crear y configurar CupertinoTextField para la edición
        text_field = CupertinoTextField(
            value=str(cell.content.value), 
            on_submit=lambda  e: self.save_edited_value(row, col, cell.content.value, page),
            on_change=lambda e: self.on_textfield_change(e, cell, page),
            autofocus=True,
            placeholder_text="",
            text_size= self.text_size,
        )

        # Actualizar el contenido de la celda para mostrar el TextField
        cell.content = text_field
        page.update()

    def on_textfield_change(self, e, cell, page):
        # Verifica si el texto comienza con '='

        if e.text.startswith("="):
            self.is_writing_formula = True
        
        if self.is_writing_formula:
            cell_references = self.parse_cell_references(cell.content.value)
            for ref in cell_references:
                cell_to_highlight = self.get_cell_from_reference(ref)
                self.iluminar(cell_to_highlight, page)
        else:
            self.is_writing_formula = False
            # Elimina el resaltado si es necesario
            # ...
       
    
    def save_edited_value(self, row, col, value, page):
        # Actualizar el valor de la celda en la estructura de datos
        self.excel_data[self.current_sheet][row][col] = value

        # Actualizar la visualización de la celda para mostrar el nuevo valor
        cell = self.cells[row][col]
        cell.content = ft.Text(value, size=self.text_size)  # Reemplazar el TextField por un Text con el nuevo valor

        self.double_clicked = False 

        
        if cell.content.value.startswith("="):
                    row, col = row, col
                    formula = cell.content.value  
                    result = Formulas().evaluate_formula(formula, "withcell", cells=self.cells,) 
                    cell.formula = formula  
                    cell.content.value = str(result) 
                    self.unhighlight_cell_colors(page)

        self.editing_cell = None  

        page.update()


    def on_pan_start(self, e: ft.DragStartEvent, page):
        print("ejecutando on_pan_start")
        for cell in self.selected_cells:
            self.unhighlight_cell(cell, page)
        self.selected_cells.clear()
        self.dragging = True
        cell = self.cells[e.control.row][e.control.col]
        self.highlight_cell(cell, page)
        self.start_cell = cell
        self.end_cell = None  # Nueva variable para almacenar la celda final


    def on_pan_update(self, e: ft.DragUpdateEvent, page):
        if not self.dragging or self.start_cell is None:
            return

        end_col = int((e.global_x) // self.cell_width) - 1
        end_row = int((e.global_y) // self.cell_height) - 3

   
        # Actualizar solo si la celda final ha cambiado
        if self.end_cell is None or end_row != self.end_cell.row or end_col != self.end_cell.col:
            self.end_cell = self.cells[end_row][end_col]

            # Identificar las nuevas celdas que deberían estar seleccionadas
            start_row, start_col = self.start_cell.row, self.start_cell.col
            end_row, end_col = self.end_cell.row, self.end_cell.col

            new_selected_cells = []

            for row in range(min(start_row, end_row), max(start_row, end_row) + 1):
                for col in range(min(start_col, end_col), max(start_col, end_col) + 1):
                    if 0 <= row < self.ROWS and 0 <= col < self.COLS:
                        new_selected_cells.append(self.cells[row][col])

            # Desresaltar las celdas que ya no están en el rango seleccionado
            for cell in self.selected_cells:
                if cell not in new_selected_cells:
                    self.unhighlight_cell(cell, page)

            # Resaltar las nuevas celdas seleccionadas
            for cell in new_selected_cells:
                if cell not in self.selected_cells:
                    self.highlight_cell(cell, page)

            self.selected_cells = new_selected_cells  # Actualizar la lista de celdas seleccionadas
       
        #print(f"Global X: {e.global_x}, Global Y: {e.global_y}")  # Imprime las coordenadas globales
        #print(f"Calculated End Col: {end_col}, Calculated End Row: {end_row}")  # Imprime las columnas y filas calculadas

    def on_pan_end(self, e, page):
        print("Ejecutando on_pan_end")
        self.dragging = False
        self.start_cell = None
        self.end_cell = None  # Limpiar la celda final
        page.update()
    
    def add_row(self, e, page):
        container_style = {
            'border': ft.border.all(0.3, ft.colors.GREEN_500),
            'border_radius': 0.2,
            'height': self.cell_height,
            'width': self.cell_width,
        }

        new_row_index = self.ROWS
        self.ROWS += 1
        new_row = []

        # Añadir una nueva fila vacía a la matriz de celdas.
        self.cells.append([None] * self.COLS)

        for c in range(self.COLS):
            tf = ft.Container(**container_style, content=Text(""))
            tf.row, tf.col = new_row_index, c
            tf.formula = None
            self.cells[new_row_index][c] = tf

            gd = ft.GestureDetector(
                mouse_cursor=ft.MouseCursor.MOVE,
                #on_pan_start=lambda e: None if self.is_mobile_device(page) or self.is_packege_device(page) else self.on_pan_start(e, page),
                #on_pan_update=lambda e: None if self.is_mobile_device(page) or self.is_packege_device(page) else self.on_pan_update(e, page),
                #on_pan_end=lambda e: None if self.is_mobile_device(page) or self.is_packege_device(page) else self.on_pan_end(e, page),
                on_tap=lambda e: self.on_single_click(e, page),
                on_double_tap=lambda e: self.on_double_click(e, page),
                #on_scroll= lambda e: self.handle_scroll_event(e, page),
            )

            gd.row, gd.col = new_row_index, c

            stacked_cell = ft.Stack(
                [
                    tf,
                    gd
                ],
                width=self.cell_width,
                height=self.cell_height
            )

            new_row.append(stacked_cell)

        # Añadir la nueva fila a las filas de la tabla
        self.table_rows.append(ft.Row(new_row, spacing=0))

        

        Row_button_style = {
            'border': ft.border.all(0.3, ft.colors.GREY),
            'border_radius': 0.2,
            'height': self.cell_height,
            'width': self.cell_height,
            'bgcolor': ft.colors.BLUE_GREY_50,
        }

        row_index_control = ft.Container(
            **Row_button_style,
            content=ft.Text(str(new_row_index + 1)),  # +1 para el índice basado en 1
            on_click= lambda e, row=new_row_index: self.on_row_index_clicked(e, page, row),
        )

        # Añadir el nuevo control de índice al final de `self.row_indices_controls`
        self.row_indices_controls.append(row_index_control)
        
        # Actualizar el control de índices de columna en la interfaz de usuario si existe
        if hasattr(self, 'row_indices'):
            self.row_indices.controls.append(row_index_control)
            self.row_indices.update()


        if self.table_initialized:  # Comprobar si la tabla se ha inicializado
            page.update()


    def add_col(self, e, page):
        container_style = {
            'border': ft.border.all(0.3, ft.colors.GREEN_500),
            'border_radius': 0.2,
            'height': self.cell_height,
            'width': self.cell_width,
        }

        new_col_index = self.COLS
        self.COLS += 1

        # Añadir una nueva columna a cada fila en la matriz de celdas
        for r in range(self.ROWS):
            tf = ft.Container(**container_style, content=ft.Text(""))
            tf.row, tf.col = r, new_col_index
            tf.formula = None
            self.cells[r].append(tf)

            gd = ft.GestureDetector(
                mouse_cursor=ft.MouseCursor.MOVE,
                #on_pan_start=lambda e: None if self.is_mobile_device(page) or self.is_packege_device(page) else self.on_pan_start(e, page),
                #on_pan_update=lambda e: None if self.is_mobile_device(page) or self.is_packege_device() else self.on_pan_update(e, page),
                #on_pan_end=lambda e: None if self.is_mobile_device(page) or self.is_packege_device() else self.on_pan_end(e, page),    
                on_tap=lambda e: self.on_single_click(e, page),
                on_double_tap=lambda e: self.on_double_click(e, page),
                #on_scroll= lambda e: self.handle_scroll_event(e, page),

            )

            gd.row, gd.col = r, new_col_index

            stacked_cell = ft.Stack(
                [
                    tf,
                    gd
                ],
                width=self.cell_width,
                height=self.cell_height
            )

            # Añadir la nueva celda al final de la fila correspondiente
            self.table_rows[r].controls.append(stacked_cell)

        # Crear el control para el índice de la nueva columna fuera del bucle for
        column_button_style = {
            'height': self.cell_height,
            'width': self.cell_width,
            'bgcolor': ft.colors.BLUE_GREY_50,
            'border': ft.border.all(0.3, ft.colors.GREY),
            'border_radius': 0.2,
        }

        new_index_label = self.num_to_excel_col(new_col_index)
        new_index_control = ft.Container(
            **column_button_style,
            content=ft.Text(new_index_label),
            on_click=lambda e, col=new_col_index: self.on_column_index_clicked(e, page, col),
        )

        # Añadir el nuevo control de índice al final de `self.column_indices_controls`
        self.column_indices_controls.append(new_index_control)

        # Actualizar el control de índices de columna en la interfaz de usuario si existe
        if hasattr(self, 'column_indices'):
            self.column_indices.controls.append(new_index_control)
            self.column_indices.update()

        if self.table_initialized:  # Comprobar si la tabla se ha inicializado
            page.update()


    
    def on_column_index_clicked(self, e, page, col_index):
        for r in range(self.ROWS):
            self.cells[r][col_index].border = ft.colors.BLUE_100
        page.update()

    def on_row_index_clicked(self, e, page, row_index):
        for c in range(self.COLS):
            self.cells[row_index][c].border = ft.colors.BLUE_100
        page.update()

    
    @staticmethod
    def num_to_excel_col(num):
        """
        Convert a zero-indexed number to a string representing its Excel column name.
        For example, 0 -> 'A', 1 -> 'B', ..., 26 -> 'AA', etc.
        """
        excel_col = ''
        while num >= 0:
            num, remainder = divmod(num, 26)
            excel_col = chr(65 + remainder) + excel_col
            num -= 1
        return excel_col
    

    def create_indices(self, page):
        # Establecemos estilos para los contenedores de índices
        special_column_button_style = {
            'height': self.cell_height,
            'width': self.cell_height,
            'bgcolor': ft.colors.BLUE_GREY_50,
            'border': ft.border.all(0.3, ft.colors.GREY),
            'border_radius': 0.2,
        }
        
        column_button_style = {
            'height': self.cell_height,
            'width': self.cell_width,
            'bgcolor': ft.colors.BLUE_GREY_50,
            'border': ft.border.all(0.3, ft.colors.GREY),
            'border_radius': 0.2,
        }
        
        row_button_style = {
            'height': self.cell_height,
            'width': 30,  # Ancho fijo para los índices de fila
            'bgcolor': ft.colors.BLUE_GREY_50,
            'border': ft.border.all(0.3, ft.colors.GREY),
            'border_radius': 0.2,
        }

        # Creamos el contenedor especial para el índice de columna separado
        special_column_container = ft.Container(
            **special_column_button_style,
            content=ft.Icon(name=ft.icons.BEACH_ACCESS, color=ft.colors.PINK),
            on_click=lambda e: self.on_special_column_clicked(e, page),
        )
        
        # Creamos los índices de columnas regulares
        column_indices_controls = [special_column_container]  # Iniciamos con el contenedor especial
        for c in range(self.COLS):
            column_label = self.num_to_excel_col(c)
            btn = ft.Container(
                **column_button_style,
                content=Text(column_label),
                on_click=lambda e, col=c: self.on_column_index_clicked(e, page, col),
            )
            column_indices_controls.append(btn)
        column_indices = ft.Row(column_indices_controls, spacing=0)

        # Creamos los índices de filas
        row_indices_controls = []
        for r in range(self.ROWS):
            btn = ft.Container(
                **row_button_style,
                content=Text(str(r + 1)),
                on_click=lambda e, row=r: self.on_row_index_clicked(e, page, row),
            )
            row_indices_controls.append(btn)
        row_indices = ft.Column(row_indices_controls, spacing=0)

        self.column_indices_controls = column_indices_controls
        self.row_indices_controls = row_indices_controls

        return column_indices, row_indices


    def update_indices(self):
        # Actualizar índices de columnas
        for c in range(self.COLS):
            column_label = self.num_to_excel_col(c + self.visible_start_col - 1)
            self.column_indices_controls[c + 1].content = Text(column_label)  # +1 porque el primer control es especial

        # Actualizar índices de filas
        for r in range(self.ROWS):
            row_label = str(r + self.visible_start_row)
            self.row_indices_controls[r].content = Text(row_label)


    def handle_scroll_event(self, e, page):
        # Calcular los índices de fila y columna basándose en el desplazamiento del scroll
        self.dragging = False

        delta_rows = int(e.scroll_delta_y / self.cell_height)
        delta_cols = int(e.scroll_delta_x / self.cell_width)

        self.visible_start_row = max(1, self.visible_start_row + delta_rows)
        self.visible_end_row = min(self.ROWS, self.visible_start_row + 12)
        
        self.visible_start_col = max(1, self.visible_start_col + delta_cols)
        self.visible_end_col = min(self.COLS, self.visible_start_col + 10)



        # Actualizar celdas visibles
        self.update_visible_cells()
        self.update_indices()


        # Actualizar la página para reflejar los cambios
        page.update()

        
    def update_visible_cells(self):
        # Verificar si se está utilizando btn_hoja y configurar el nombre de la hoja
        sheet_name = self.current_sheet if not self.btn_hoja else self.current_sheet

        if sheet_name not in self.edited_cells:
            self.edited_cells[sheet_name] = {}

        if self.edited_cells:
            self.save_excel_data()

        for r in range(self.ROWS):
            for c in range(self.COLS):
                # Calcular la posición real de la celda en los datos
                data_row = r + self.visible_start_row - 1
                data_col = c + self.visible_start_col - 1

                # Comprobar si la celda ha sido editada
                if (data_row, data_col) in self.edited_cells[sheet_name]:
                    cell_value = self.edited_cells[sheet_name][(data_row, data_col)]
                    
                else:
                    # Si no ha sido editada, usar el valor de la hoja de cálculo
                    if sheet_name in self.excel_data and \
                    data_row < len(self.excel_data[sheet_name]) and \
                    data_col < len(self.excel_data[sheet_name][data_row]):
                        cell_value = self.excel_data[sheet_name][data_row][data_col]
                        if cell_value is None:
                            cell_value = ""
                    else:
                        cell_value = ""

                # Comprobar si la celda contiene una fórmula
                if isinstance(cell_value, str) and cell_value.startswith("="):
                    # Evaluar la fórmula y actualizar el valor de la celda
                    try:
                        evaluated_value = Formulas().evaluate_formula(cell_value, "withexceldata", excel_data= self.excel_data, current_sheet_name= self.current_sheet)
                        cell_display_value = str(evaluated_value)
                    except Exception as e:
                        cell_display_value = "Error"
                else:
                    # Si no es una fórmula, mostrar el valor como está
                    cell_display_value = str(cell_value) if cell_value != "" else ""

                # Actualizar el valor de la celda en la interfaz de usuario
                self.cells[r][c].content.value = cell_display_value
                            
    def on_horizontal_slider_change(self, e, page):
        # Calcula el desplazamiento en las columnas basado en el valor del slider
        self.visible_start_col = int(e.control.value)
        self.visible_end_col = self.visible_end_col
        self.update_visible_cells()
        self.update_indices()
        page.update()


    def on_vertical_slider_change(self, e, page):
        # Calcula el desplazamiento en las filas basado en el valor del slider
        self.visible_start_row = int(e.control.value)
        self.visible_end_row = self.visible_end_row
        self.update_visible_cells()
        self.update_indices()
        page.update()

    # Crear slider horizontal
    def horizontal_slider(self, page):
        slider = ft.Container(
                ft.CupertinoSlider(
                min=1, 
                max=10000, 
                on_change=lambda e: self.on_horizontal_slider_change(e, page),
                active_color=ft.colors.GREEN_500,
                thumb_color=ft.colors.GREY_50,
                ),
                width= page.width * 0.9

        )
        return slider

    # Crear slider vertical
    def vertical_slider(self, page):
        slider = ft.Container(
                    ft.CupertinoSlider(
                        min=1, 
                        max=10000, 
                        on_change=lambda e: self.on_vertical_slider_change(e, page),
                        active_color=ft.colors.GREEN_500,
                        thumb_color=ft.colors.GREY_50,
                        rotate= 1.57079632679,
                    ),
                    #alignment= alignment.top_left,
                    #width= page.height * 0.8,
                    height= page.height *0.7
        )  
        return slider

    def create_sheets_section(self, page):
        """
        Crea una sección con botones para cada hoja de Excel.
        """
        buttons = []
        for sheet_name in self.excel_data.keys():
            btn = ft.ElevatedButton(
                    adaptive=True, 
                    bgcolor=ft.colors.BLUE_GREY_50,
                    on_click= lambda e, name=sheet_name: self.on_sheet_selected(e, page, name),
                    content=ft.Row(
                        [
                            ft.Icon(name=ft.cupertino_icons.CIRCLE_GRID_3X3, color="pink"),
                            ft.Text(sheet_name, color= ft.colors.BLACK, size=self.text_size),
                        ],
                        tight=True,
                       
                    ),
            )
            
           

            buttons.append(btn)
        return ft.Row(buttons)

    def on_sheet_selected(self, e, page, sheet_name):
        """
        Maneja la selección de una hoja de Excel.
        """
        self.current_sheet = sheet_name
        self.btn_hoja = True
        self.update_visible_cells()  # Asegúrate de que esta función use self.current_sheet
        page.update()
        self.btn_hoja = False

    def integrate_edits_to_excel_data(self):
        """
        Integra los cambios editados directamente en el archivo Excel.
        """
        workbook = openpyxl.load_workbook(self.excel_file_path)

        for sheet_name, edits in self.edited_cells.items():
            if sheet_name not in workbook.sheetnames:
                workbook.create_sheet(sheet_name)
            worksheet = workbook[sheet_name]

            for (row, col), value in edits.items():
                cell = worksheet.cell(row=row + 1, column=col + 1)
                if isinstance(value, str) and value.startswith('='):
                    # Eliminar '=' adicionales al principio
                    while value.startswith('=='):
                        value = value[1:]
                    print(f"Guardando fórmula en celda ({row + 1}, {col + 1}): {value}")
                    cell.value = value
                else:
                    cell.value = value

        workbook.save(self.excel_file_path)
        print(f"Datos guardados exitosamente en {self.excel_file_path}")

    def save_excel_data(self):
        """
        Guarda los cambios actualizados en el archivo Excel.
        """
        self.integrate_edits_to_excel_data()


    def integrate_edits_to_all_excel_data(self):
        """
        Integra los cambios editados en la estructura de datos principal de Excel antes de guardar.
        """
        for sheet_name, edits in self.edited_cells.items():
            for (row, col), value in edits.items():
                # Asegúrate de que la hoja y las filas/columnas existen en excel_data
                if sheet_name not in self.excel_data:
                    self.excel_data[sheet_name] = []
                while len(self.excel_data[sheet_name]) <= row:
                    self.excel_data[sheet_name].append([])
                while len(self.excel_data[sheet_name][row]) <= col:
                    self.excel_data[sheet_name][row].append(None)
                # Aplicar el valor editado
                self.excel_data[sheet_name][row][col] = value

    def save_all_excel_data(self):
        """
        Guarda los datos actualizados en el archivo Excel.
        """
        self.integrate_edits_to_excel_data()

        workbook = openpyxl.load_workbook(self.excel_file_path)

        # Iterar sobre todas las hojas en self.excel_data
        for sheet_name, sheet_data in self.excel_data.items():
            if sheet_name not in workbook.sheetnames:
                workbook.create_sheet(sheet_name)
            worksheet = workbook[sheet_name]

            for row_index, row in enumerate(sheet_data):
                for col_index, value in enumerate(row):
                    cell = worksheet.cell(row=row_index + 1, column=col_index + 1)
                    if isinstance(value, str) and value.startswith('='):
                        # Eliminar '=' adicionales al principio
                        while value.startswith('=='):
                            value = value[1:]
                        print(f"Guardando fórmula en celda ({row_index + 1}, {col_index + 1}): {value}")
                        cell.value = value
                    else:
                        cell.value = value

        workbook.save(self.excel_file_path)
        print(f"Datos guardados exitosamente en {self.excel_file_path}")
    
    def format_date(self, date_input):
        if isinstance(date_input, datetime.datetime):
            # El objeto ya es datetime, por lo que se puede formatear directamente
            return date_input.strftime('%d-%B-%Y')
        elif isinstance(date_input, str):
            # Convertir la cadena a un objeto datetime
            date_obj = datetime.datetime.strptime(date_input, '%Y-%m-%d %H:%M:%S')
            return date_obj.strftime('%d-%B-%Y')
        else:
            # No es una fecha, devuelve el valor original o una cadena vacía
            return str(date_input) 

    def format_number(self, value):
        """Formatea un número con comas como separadores de miles."""
        return "{:,}".format(value).replace(",", ".")

    def create_table(self, page):
        page.on_keyboard_event = lambda e: self.on_keyboard_event(e, page)
        self.create_formula_container(page)  # Inicializar el contenedor de fórmulas


        container_style = {
            'border': ft.border.all(0.3, ft.colors.GREEN_500),
            'border_radius': 0.2,
            'height': self.cell_height,
            'width': self.cell_width,
        }

        

        #crear los indices de las celdas
        column_indices, row_indices = self.create_indices(page)

        
        #crear filas y columnas para la tabla usando bucles
        sheet_name = self.current_sheet
        self.table_rows= []
        for r in range(self.ROWS):
            row_cells = []
            for c in range(self.COLS):
                cell_value = self.excel_data[sheet_name][r][c] if r < len(self.excel_data[sheet_name]) and c < len(self.excel_data[sheet_name][r]) else ""
                cell_content = ""                
                
                # Verificar si la celda tiene una fórmula y calcularla
                if isinstance(cell_value, str) and cell_value.startswith('='):
                    cell_content = str(Formulas().evaluate_formula(cell_value, "withexceldata", excel_data= self.excel_data, current_sheet_name=self.current_sheet))
                
                elif cell_value:
                    try:
                        # Intenta formatear como fecha
                        cell_content = self.format_date(cell_value)
                    except ValueError:
                        # Si no es una fecha, mantiene el valor original
                        cell_content = cell_value
                    
                
               
                custom_container_style = container_style.copy()
                custom_container_style['bgcolor'] = "#FFFFFF" if r % 2 == 0 else "#EEEEEE"
                
                tf = ft.Container(**custom_container_style, content=ft.Text(cell_content, size=self.text_size)) 


                tf.row, tf.col = r, c
                tf.formula = cell_value
                self.cells[r][c] = tf

                gd = ft.GestureDetector(
                    mouse_cursor=ft.MouseCursor.MOVE,
                    #on_pan_start=lambda e: None if self.is_mobile_device(page) or self.is_packege_device(page) else self.on_pan_start(e, page),
                    #on_pan_update=lambda e: None if self.is_mobile_device(page) or self.is_packege_device(page) else self.on_pan_update(e, page),
                    #on_pan_end=lambda e: None if self.is_mobile_device(page) or self.is_packege_device(page) else self.on_pan_end(e, page),
                    on_tap=lambda e: self.on_single_click(e, page),
                    on_double_tap=lambda e: self.on_double_click(e, page),
                    on_scroll= lambda e: self.handle_scroll_event(e, page),
                )

                gd.row, gd.col = r, c
                
                # Aquí está el Stack
                stacked_cell = ft.Stack(
                    [
                        tf,  # TextField en la parte inferior
                        gd   # GestureDetector en la parte superior
                    ],
                    width=self.cell_width,
                    height=self.cell_height
                )
                
                row_cells.append(stacked_cell)  # Añadir el Stack en lugar del GestureDetector

             
            #añadir nueva fila a la lista de filas
            self.table_rows.append(ft.Row(row_cells, spacing=0))
        
        
        # Crear una columna con todas las filas para permitir desplazamiento vertical
        table_column = ft.Column(self.table_rows,  spacing=0, scroll=ft.ScrollMode.ALWAYS)

        # Envolver la columna en un contenedor Row para desplazamiento horizontal
        scrollable_columns = ft.Row([row_indices, table_column], spacing=0)

        # Configurar el evento de scroll en el contenedor que quieres que sea desplazable

        #scrollable_columns.on_scroll = self.handle_scroll_event

        #indices de filas
        tabla_indices = ft.Column(
            [column_indices, scrollable_columns],
            spacing=0,
            width= page.width * 0.90 ,
        )

        seccion_hojas = self.create_sheets_section(page)
        vertical_slider = self.vertical_slider(page)
        horizontal_slider = self.horizontal_slider(page)

        final_table = ft.Column([
            ft.Row([
                tabla_indices,
                vertical_slider
            ],
                spacing=0,
                height= page.height * 0.70,
                scroll=ft.ScrollMode.HIDDEN),
            ft.Row([
                seccion_hojas,
            ],
                spacing=0,
                height= page.height *0.05),
            ft.Row([
                horizontal_slider
            ],  
                spacing=0,
                height= page.height *0.05)
           
        ],
        spacing=0.6,
        scroll=ft.ScrollMode.HIDDEN,
        expand=True)
        

        final_layout = ft.Column([
            self.formula_container,  # Añadir el contenedor de fórmulas en la parte superior
            final_table  
        ],
        spacing=0.7)
        return final_layout