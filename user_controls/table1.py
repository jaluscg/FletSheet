import flet as ft
from flet import *
import re
from fletsheet_formula import Formulas
import openpyxl
import sys 
import os
import random
import datetime


class TextFieldTable():
    """
    Una matriz de celdas que se puede editar y desplazar.
    """

    def __init__(self, 
                excel_file_path,
                page_width, 
                page_height,
                cell_height: OptionalNumber = 30,
                cell_width: OptionalNumber = 100,
                ):
        
        self.page_width = page_width
        self.page_height = page_height
        self.ROWS = int((page_height / 30)*0.8)
        self.COLS = int((page_width / 100)*0.9)
        self.selected_cells = [] #inicializar como lista vacía
        self.cells =  [[None for _ in range(self.COLS)] for _ in range(self.ROWS)]  # Matriz de celdas
        self.dragging = False
        self.double_clicked = False
        self.current_selected_cell = None
        self.editing_cell = None
        self.clipboard = [] #contenido copiado o cortado puede ser una lista
        self.double_clicked = False
        self.visible_start_row = 1
        self.visible_end_row =  self.ROWS
        self.visible_start_col = 1
        self.visible_end_col = self.COLS
        self.start_cell = None 
        self.table_rows = []
        self.table_initialized = False  # Inicializa el estado de la tabla
        self.cell_height = cell_height  
        self.cell_width = cell_width
        self.btn_hoja = False
        self.edited_cells = {}  
        self.undo_stack = []  # Pila para deshacer cambios
        self.excel_file_path = excel_file_path
        self.excel_data = self.load_excel_data(self.excel_file_path)
        self.current_sheet = next(iter(self.excel_data), None)
        self.text_size = 14
        self.is_writing_formula = False
        self.cell_colors = {}  # Diccionario para almacenar los colores de las celdas
        self.formula_container = None
        self.container_row = None
        self.container_col = None

 
    def load_excel_data(self, filepath):
        print("se está cargando data excel")
        workbook = openpyxl.load_workbook(filepath, data_only=False)  # Cambiar data_only a False
        data = {}

        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            sheet_data = []
            for row in worksheet.iter_rows():
                row_data = []
                for cell in row:
                    if cell.data_type == 'f':  # Verificar si la celda contiene una fórmula
                        cell_value = cell.value  # Preceder la fórmula con '='
                    else:
                        cell_value = cell.value  # Obtener el valor de la celda como es
                    row_data.append(cell_value)
                sheet_data.append(row_data)
            data[sheet_name] = sheet_data

        return data

    
    def create_table(self, page):
        page.on_keyboard_event = lambda e: self.on_keyboard_event(e, page)
        self.create_formula_container(page)  # Inicializar el contenedor de fórmulas


        container_style = {
            'border': ft.border.all(0.3, ft.colors.GREEN_500),
            'border_radius': 0.2,
            'height': self.cell_height,
            'width': self.cell_width,
        }

        

        #crear los indices de las celdas
        column_indices, row_indices = self.create_indices(page)

        
        #crear filas y columnas para la tabla usando bucles
        sheet_name = self.current_sheet
        self.table_rows= []
        for r in range(self.ROWS):
            row_cells = []
            for c in range(self.COLS):
                cell_value = self.excel_data[sheet_name][r][c] if r < len(self.excel_data[sheet_name]) and c < len(self.excel_data[sheet_name][r]) else ""
                cell_content = ""                
                
                # Verificar si la celda tiene una fórmula y calcularla
                if isinstance(cell_value, str) and cell_value.startswith('='):
                    cell_content = str(Formulas().evaluate_formula(self.cells, cell_value, r, c, "withexceldata", self.excel_data[sheet_name]))
                
                elif cell_value:
                    try:
                        # Intenta formatear como fecha
                        cell_content = self.format_date(cell_value)
                    except ValueError:
                        # Si no es una fecha, mantiene el valor original
                        cell_content = cell_value
                    
                
               
                custom_container_style = container_style.copy()
                custom_container_style['bgcolor'] = "#FFFFFF" if r % 2 == 0 else "#EEEEEE"
                
                tf = ft.Container(**custom_container_style, content=ft.Text(cell_content, size=self.text_size)) 


                tf.row, tf.col = r, c
                tf.formula = cell_value
                self.cells[r][c] = tf

           